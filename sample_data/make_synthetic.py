"""
Build a synthetic test workbook based on the SAP template, with deliberately
bad and good rows to exercise every validator rule.
"""
import shutil
from pathlib import Path

import xlrd
from openpyxl import load_workbook  # write to .xlsx since xlwt is dead


SOURCE = Path(__file__).parent / "Product Master Creation.xls"
TARGET_XLSX = Path(__file__).parent / "Product Master Test Data.xlsx"


def main():
    # convert .xls -> .xlsx by replaying cells (xlrd reads, openpyxl writes)
    src_book = xlrd.open_workbook(str(SOURCE))

    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)

    for sname in src_book.sheet_names():
        sh = src_book.sheet_by_name(sname)
        ws = wb.create_sheet(title=sname[:31])  # excel sheet name limit
        for r in range(sh.nrows):
            for c in range(sh.ncols):
                v = sh.cell_value(r, c)
                if v == "":
                    continue
                ws.cell(row=r+1, column=c+1, value=v)

    # ----- Inject test rows -----

    bd = wb["Basic Data"]
    # Header row 4 (1-based 5) has SAP fields. Data starts at row 9 (index 8).
    # SAP fields at row 5 (1-based) - we know columns:
    # PRODUCT(1) MTART(2) ATTYP(3) SATNR(4) MATKL(5) MBRSH(6) MAKTX(7) SPRAS(8) MEINS(9)
    # AENNR(10), then more later: MHDHB column - need to find it
    sap_fields = [bd.cell(row=5, column=c).value for c in range(1, bd.max_column + 1)]
    def col(name):
        return sap_fields.index(name) + 1 if name in sap_fields else None

    rows = [
        # Row 9: Good finished product (FERT) measured in PC
        dict(PRODUCT="MAT-1001", MTART="FERT", MAKTX="Wireless Mouse", SPRAS="EN", MEINS="PC"),
        # Row 10: Oil with PC unit (UoM mismatch)
        dict(PRODUCT="MAT-1002", MTART="ROH", MAKTX="Engine Oil 5W-30", SPRAS="EN", MEINS="PC"),
        # Row 11: Cable with EA unit (length expected — flagged)
        dict(PRODUCT="MAT-1003", MTART="HALB", MAKTX="Copper Cable 2.5mm", SPRAS="EN", MEINS="EA"),
        # Row 12: Description = oil but no MTART (mandatory missing)
        dict(PRODUCT="MAT-1004", MAKTX="Olive Oil Extra Virgin", SPRAS="EN", MEINS="L"),
        # Row 13: Service with non-zero text in unsupported field — too-long description
        dict(PRODUCT="MAT-1005", MTART="DIEN", MAKTX="A" * 60, SPRAS="EN", MEINS="HR"),
        # Row 14: Perishable food without shelf life
        dict(PRODUCT="MAT-1006", MTART="HAWA", MAKTX="Fresh Whole Milk 1L", SPRAS="EN", MEINS="L"),
        # Row 15: Description with forbidden term for material type
        dict(PRODUCT="MAT-1007", MTART="VERP", MAKTX="Finished Steel Pump", SPRAS="EN", MEINS="PC"),
        # Row 16: missing PRODUCT (mandatory) and bad SPRAS — leave product blank by giving description only
        dict(MTART="ROH", MAKTX="Raw Iron Ore", MEINS="KG"),
    ]

    start_row = 9  # 1-based — first empty data row
    for i, r in enumerate(rows):
        for k, v in r.items():
            c = col(k)
            if c:
                bd.cell(row=start_row + i, column=c, value=v)

    # ----- Plant Data -----
    pd_ws = wb["Plant Data"]
    pd_fields = [pd_ws.cell(row=5, column=c).value for c in range(1, pd_ws.max_column + 1)]
    def pcol(name):
        return pd_fields.index(name) + 1 if name in pd_fields else None
    pd_rows = [
        # MRP type set, controller missing, lot size missing
        dict(PRODUCT="MAT-1001", WERKS="P001", DISMM="PD"),
        # VB reorder type, no reorder point
        dict(PRODUCT="MAT-1002", WERKS="P001", DISMM="VB", DISPO="100", DISLS="EX"),
        # Negative rounding
        dict(PRODUCT="MAT-1003", WERKS="P001", BSTRF=-5),
        # External procurement no purchasing group
        dict(PRODUCT="MAT-1006", WERKS="P002", BESKZ="F"),
        # Profit center but no controlling area
        dict(PRODUCT="MAT-1001", WERKS="P002", PRCTR="PC100"),
        # Plant status w/o valid from
        dict(PRODUCT="MAT-1005", WERKS="P002", MMSTA="01"),
        # Reorder>Max
        dict(PRODUCT="MAT-1007", WERKS="P003", DISMM="VB", DISPO="200", DISLS="EX", MINBE=100, MABST=50),
    ]
    for i, r in enumerate(pd_rows):
        for k, v in r.items():
            c = pcol(k)
            if c:
                pd_ws.cell(row=9 + i, column=c, value=v)

    # ----- Valuation Data -----
    val = wb["Valuation Data"]
    val_fields = [val.cell(row=5, column=c).value for c in range(1, val.max_column + 1)]
    def vcol(n):
        return val_fields.index(n) + 1 if n in val_fields else None
    val_rows = [
        # Standard price control with no price
        dict(PRODUCT="MAT-1001", BWKEY="P001", BKLAS="3000", VPRSV="S"),
        # V control no moving price
        dict(PRODUCT="MAT-1002", BWKEY="P001", BKLAS="3001", VPRSV="V"),
        # Negative price
        dict(PRODUCT="MAT-1003", BWKEY="P001", BKLAS="3001", VPRSV="S", STPRS=-12.5),
        # No valuation class
        dict(PRODUCT="MAT-1004", BWKEY="P001", VPRSV="S", STPRS=10),
        # Bad price control letter
        dict(PRODUCT="MAT-1005", BWKEY="P002", BKLAS="3200", VPRSV="X", STPRS=100),
        # Out-of-band high price for VERP packaging
        dict(PRODUCT="MAT-1007", BWKEY="P002", BKLAS="3010", VPRSV="S", STPRS=99999),
        # Reference unknown product (not in Basic Data)
        dict(PRODUCT="MAT-9999", BWKEY="P001", BKLAS="3000", VPRSV="S", STPRS=50),
    ]
    for i, r in enumerate(val_rows):
        for k, v in r.items():
            c = vcol(k)
            if c:
                val.cell(row=9 + i, column=c, value=v)

    # ----- Alternative UoM -----
    alt = wb["Alternative Units of Measure"]
    alt_fields = [alt.cell(row=5, column=c).value for c in range(1, alt.max_column + 1)]
    def acol(n):
        return alt_fields.index(n) + 1 if n in alt_fields else None
    alt_rows = [
        # A) equal to base UoM (PC)
        dict(PRODUCT="MAT-1001", MEINH="PC", UMREN=1, UMREZ=1),
        # B) missing numerator/denominator
        dict(PRODUCT="MAT-1001", MEINH="BOX"),
        # C) duplicate alt UoM BOX
        dict(PRODUCT="MAT-1001", MEINH="BOX", UMREN=1, UMREZ=10),
        # D) GTIN bad length, dimensions without unit
        dict(PRODUCT="MAT-1002", MEINH="CAR", UMREN=1, UMREZ=12, EAN11="12345", LAENG=10),
        # E) impossible: base=PC (count), alt=KG (mass) — cross-dim warning, but not physically impossible
        dict(PRODUCT="MAT-1001", MEINH="KG", UMREN=1, UMREZ=1),
        # F) magnitude error: 1 KG = 10 G (should be 1000 G)
        dict(PRODUCT="MAT-1002", MEINH="G", UMREN=10, UMREZ=1),
        # G) physically impossible: base=L (volume), alt=M (length)
        dict(PRODUCT="MAT-1006", MEINH="M", UMREN=1, UMREZ=1),
        # H) implausible same-dim ratio: 1 KG base → 1e8 alt KG  (impossible)
        dict(PRODUCT="MAT-1002", MEINH="KG", UMREN=100000000, UMREZ=1),
        # I) trivial 1:1 ratio between different same-dimension units (L→ML should be 1000:1)
        dict(PRODUCT="MAT-1006", MEINH="ML", UMREN=1, UMREZ=1),
        # J) service material (DIEN) with a mass alt UoM — contextual mismatch
        dict(PRODUCT="MAT-1005", MEINH="KG", UMREN=1, UMREZ=1),
        # K) good cable (M base) → CM with correct 100:1 ratio
        dict(PRODUCT="MAT-1003", MEINH="CM", UMREN=100, UMREZ=1, EAN11="0123456789012"),
    ]
    for i, r in enumerate(alt_rows):
        for k, v in r.items():
            c = acol(k)
            if c:
                alt.cell(row=9 + i, column=c, value=v)

    # ----- Distribution Chains (sales) -----
    dc = wb["Distribution Chains"]
    dc_fields = [dc.cell(row=5, column=c).value for c in range(1, dc.max_column + 1)]
    def dcol(n):
        return dc_fields.index(n) + 1 if n in dc_fields else None
    dc_rows = [
        # missing sales org
        dict(PRODUCT="MAT-1001"),
        # status without valid from
        dict(PRODUCT="MAT-1002", VKORG="1000", VTWEG="10", VMSTA="01"),
        # negative min order
        dict(PRODUCT="MAT-1003", VKORG="1000", VTWEG="10", AUMNG=-1),
        # item cat group without acct assignment
        dict(PRODUCT="MAT-1004", VKORG="1000", VTWEG="10", MTPOS="NORM"),
    ]
    for i, r in enumerate(dc_rows):
        for k, v in r.items():
            c = dcol(k)
            if c:
                dc.cell(row=9 + i, column=c, value=v)

    # ----- Storage Locations -----
    sl = wb["Storage Locations"]
    sl_fields = [sl.cell(row=5, column=c).value for c in range(1, sl.max_column + 1)]
    def scol(n):
        return sl_fields.index(n) + 1 if n in sl_fields else None
    sl_rows = [
        # plant not in plant data
        dict(PRODUCT="MAT-1001", WERKS="ZZ99", LGORT="0001"),
        # missing storage location key
        dict(PRODUCT="MAT-1002", WERKS="P001"),
    ]
    for i, r in enumerate(sl_rows):
        for k, v in r.items():
            c = scol(k)
            if c:
                sl.cell(row=9 + i, column=c, value=v)

    wb.save(TARGET_XLSX)
    print(f"Wrote {TARGET_XLSX}")


if __name__ == "__main__":
    main()
