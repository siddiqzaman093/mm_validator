"""
SAP UOM master data loader.

Reads SAP_UOM_All.xlsx (UOM sheet) and builds authoritative lookup tables used
by all UoM validation modules.

ISO codes (column B) are the PRIMARY lookup key.
SAP internal codes (column A) are the FALLBACK for rows that have no ISO code.

Exposed constants
-----------------
ISO_DIM      : dict[str, str]   lowercase_iso_code  -> dimension_category string
SAP_DIM      : dict[str, str]   lowercase_sap_code  -> dimension_category string
ISO_INFO     : dict[str, dict]  lowercase_iso_code  -> {sap, iso, desc, dim_id, dim_desc, category}
SAP_INFO     : dict[str, dict]  lowercase_sap_code  -> same
VALID_ISO    : set[str]         all known lowercase ISO codes
VALID_SAP    : set[str]         all known lowercase SAP codes

Dimension categories used
-------------------------
  mass | volume | length | area | time | count | packaging |
  mole | energy | pressure | temperature | proportion | density |
  speed | force | power | other
"""
from __future__ import annotations

import os
from pathlib import Path


# ---------------------------------------------------------------------------
# AAAADL ("no dimensions") sub-classification
# ---------------------------------------------------------------------------
_PACKAGING_SAP = {
    "bag", "bt", "box", "can", "car", "crt", "cv", "dr",
    "kit", "pac", "pal", "rol", "set",
}
_COUNT_SAP = {
    "au", "cop", "dz", "ea", "gro", "paa", "pc", "prc",
    "prs", "teu", "ts",
}

# Corresponding ISO codes for packaging / count UOMs
_PACKAGING_ISO = {
    "bg",  # BAG
    "bo",  # BT (bottle)
    "bx",  # BOX
    "ca",  # CAN (canister)
    "ct",  # CAR (carton)
    "cr",  # CRT (crate)
    "cs",  # CV (case)
    "dr",  # DR (drum)
    "kt",  # KIT
    "pk",  # PAC (pack)
    "pf",  # PAL (pallet)
    "ro",  # ROL (roll)
}
_COUNT_ISO = {
    "c62",  # AU (activity unit)
    "dzn",  # DZ (dozen)
    "ea",   # EA (each)
    "gro",  # GRO (gross)
    "pr",   # PAA (pair)
    "pce",  # PC (piece) — ISO 31 code
}

# SAP dimension-ID -> simplified category used by validators
_DIM_ID_CATEGORY: dict[str, str] = {
    "MASS":   "mass",
    "VOLUME": "volume",
    "LENGTH": "length",
    "SURFAC": "area",
    "TIME":   "time",
    "MOLQU":  "mole",
    "ENERGY": "energy",
    "PRESS":  "pressure",
    "TEMP":   "temperature",
    "PROPOR": "proportion",
    "MPROPO": "proportion",
    "VPROPO": "proportion",
    "DENSI":  "density",
    "SPEED":  "speed",
    "FORCE":  "force",
    "POWER":  "power",
    # AAAADL is handled per-code below
}


def _find_uom_file() -> Path | None:
    """Search for SAP_UOM_All.xlsx relative to this module's directory tree."""
    base = Path(__file__).parent
    candidates = [
        base.parent / "sample_data" / "SAP_UOM_All.xlsx",
        base.parent / "data" / "SAP_UOM_All.xlsx",
        base.parent / "SAP_UOM_All.xlsx",
    ]
    for p in candidates:
        if p.exists():
            return p
    return None


def _load() -> tuple[dict, dict, dict, dict]:
    """Return (SAP_DIM, SAP_INFO, ISO_DIM, ISO_INFO)."""
    uom_file = _find_uom_file()
    if uom_file is None:
        return {}, {}, {}, {}

    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(uom_file), data_only=True, read_only=True)
        ws = wb["UOM"]
    except Exception:
        return {}, {}, {}, {}

    sap_dim: dict[str, str] = {}
    sap_info: dict[str, dict] = {}
    iso_dim: dict[str, str] = {}
    iso_info: dict[str, dict] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        sap_code = str(row[0]).strip()
        iso       = str(row[1]).strip() if row[1] else ""
        desc      = str(row[2]).strip() if row[2] else ""
        dim_id    = str(row[3]).strip() if row[3] else ""
        dim_desc  = str(row[4]).strip() if row[4] else ""

        sap_key = sap_code.lower()
        if not sap_key:
            continue

        # Determine category
        if dim_id == "AAAADL":
            if sap_key in _PACKAGING_SAP:
                cat = "packaging"
            elif sap_key in _COUNT_SAP:
                cat = "count"
            else:
                cat = "other"
        else:
            cat = _DIM_ID_CATEGORY.get(dim_id, "other")

        entry = {
            "sap": sap_code,
            "iso": iso,
            "desc": desc,
            "dim_id": dim_id,
            "dim_desc": dim_desc,
            "category": cat,
        }

        sap_dim[sap_key] = cat
        sap_info[sap_key] = entry

        # Build ISO lookup (primary)
        if iso:
            iso_key = iso.lower()
            # Refine AAAADL category using ISO sets
            if dim_id == "AAAADL":
                if iso_key in _PACKAGING_ISO:
                    cat = "packaging"
                elif iso_key in _COUNT_ISO:
                    cat = "count"
                # else: keep whatever was derived from SAP key above
            iso_dim[iso_key] = cat
            iso_info[iso_key] = {**entry, "category": cat}

    return sap_dim, sap_info, iso_dim, iso_info


SAP_DIM, SAP_INFO, ISO_DIM, ISO_INFO = _load()
VALID_SAP: set[str] = set(SAP_DIM.keys())
VALID_ISO: set[str] = set(ISO_DIM.keys())


def classify_sap_uom(code: str) -> str | None:
    """
    Return dimension category for a UOM code.
    Checks ISO codes first (priority), then SAP internal codes as fallback.
    Returns None if the code is not recognised in either table.
    """
    if not code:
        return None
    key = code.strip().lower()
    return ISO_DIM.get(key) or SAP_DIM.get(key)


def describe_sap_uom(code: str) -> str:
    """Return a human description like 'Kilogram (mass)'. ISO lookup first."""
    if not code:
        return ""
    key = code.strip().lower()
    info = ISO_INFO.get(key) or SAP_INFO.get(key)
    if not info:
        return f"'{code}' (unknown UOM)"
    return f"{info['desc']} ({info['category']})"


def is_valid_sap_uom(code: str) -> bool:
    """True if code is a recognised ISO code OR a recognised SAP internal code."""
    if not code:
        return False
    key = code.strip().lower()
    return key in VALID_ISO or key in VALID_SAP
