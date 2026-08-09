"""
Load SAP S/4HANA Migration Cockpit `Product Master Creation.xls` template.

Layout per data sheet:
  row 0: title
  row 1: version / copyright
  row 2: blank
  row 3: SAP structure (e.g. S_MARA, S_MARC, S_MARM, S_MBEW, S_MVKE)
  row 4: SAP field codes (e.g. PRODUCT, MTART, MEINS, ...)
  row 5: format string ETE;80;0;C;80;0
  row 6: group label
  row 7: long description (field name + help text + Type/Length)
  row 8+: actual data rows

Field List sheet header at row 3, values from row 4 onward.
Sheet Name column (col B = idx 1) holds either:
  - "Basic Data (mandatory)" / "Plant Data (optional)" -> sheet header
  - blank -> field row belonging to last seen sheet header

Sheets are read in ONE sequential pass each (openpyxl read_only streaming for
.xlsx) and stored in the compact Row structure from models.py — a 60k-row
workbook must fit comfortably in a 512MB container, which rules out both
openpyxl's full cell-object model and a dict per cell.
"""
from __future__ import annotations

import io
import re
from typing import Any, Iterator

import xlrd

from .models import FieldSpec, Row, SheetData


class _Book:
    """Uniform wrapper over an xlrd or openpyxl(read-only) workbook that only
    exposes what the loader needs: sheet names and a sequential row iterator."""

    def __init__(self, raw, kind: str):
        self._raw = raw
        self.kind = kind  # "xlrd" or "openpyxl"

    def sheet_names(self) -> list[str]:
        if self.kind == "xlrd":
            return self._raw.sheet_names()
        return list(self._raw.sheetnames)

    def iter_sheet(self, name: str) -> Iterator[tuple[list, list | None]]:
        """Yield (values, cell_types) per row. cell_types is None for .xlsx —
        types are derived from the Python values; xlrd needs real types because
        it reads dates as plain floats."""
        if self.kind == "xlrd":
            sh = self._raw.sheet_by_name(name)
            for r in range(sh.nrows):
                yield sh.row_values(r), sh.row_types(r)
        else:
            ws = self._raw[name]
            for row in ws.iter_rows(values_only=True):
                yield list(row), None

    def close(self) -> None:
        if self.kind == "openpyxl":
            try:
                self._raw.close()
            except Exception:
                pass


HEADER_ROW_SAP_FIELDS = 4
HEADER_ROW_FORMAT = 5
HEADER_ROW_DESC = 7
DATA_START_ROW = 8


def _clean(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    return s


def _to_int(value) -> int | None:
    s = _clean(value)
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def _at(values: list, i: int) -> Any:
    """Index with tolerance — streamed rows can be shorter than the header."""
    return values[i] if i < len(values) else None


def _strip_sheet_suffix(name: str) -> str:
    """ 'Basic Data (mandatory)' -> 'Basic Data' """
    return re.sub(r"\s*\((mandatory|optional)\)\s*$", "", name, flags=re.I).strip()


def load_field_list(book: _Book) -> dict[tuple[str, str], FieldSpec]:
    """Return mapping (sheet_name, sap_field) -> FieldSpec."""
    if "Field List" not in book.sheet_names():
        return {}   # Field List sheet absent — validation continues without it
    specs: dict[tuple[str, str], FieldSpec] = {}
    current_sheet: str | None = None

    for r, (values, _types) in enumerate(book.iter_sheet("Field List")):
        if r < 4:
            continue
        sheet_cell = _clean(_at(values, 1))
        if sheet_cell:
            current_sheet = _strip_sheet_suffix(sheet_cell)
            continue
        if not current_sheet:
            continue

        group = _clean(_at(values, 2))
        desc = _clean(_at(values, 3))
        importance = _clean(_at(values, 4))
        ftype = _clean(_at(values, 5))
        length = _to_int(_at(values, 6))
        decimal = _to_int(_at(values, 7))
        sap_struct = _clean(_at(values, 8))
        sap_field = _clean(_at(values, 9))

        if not desc and not sap_field:
            continue

        specs[(current_sheet, sap_field)] = FieldSpec(
            sheet=current_sheet,
            group=group,
            description=desc,
            importance=importance,
            type=ftype,
            length=length,
            decimal=decimal,
            sap_structure=sap_struct,
            sap_field=sap_field,
        )
    return specs


def load_sheet_data(book: _Book, sheet_name: str) -> SheetData | None:
    if sheet_name not in book.sheet_names():
        return None

    sap_structure = ""
    sap_fields: list[str] = []
    descriptions: list[str] = []
    # sap_field -> position in each row's compact value tuple; SHARED by every
    # Row of this sheet.
    keymap: dict[str, int] = {}
    cols: list[int] = []          # source column index per keymap entry
    rows: list[Row] = []

    for r, (values, types) in enumerate(book.iter_sheet(sheet_name)):
        if r == 3:
            sap_structure = _clean(_at(values, 0))
            continue
        if r == HEADER_ROW_SAP_FIELDS:
            sap_fields = [_clean(v) for v in values]
            for c, sap_field in enumerate(sap_fields):
                if sap_field:
                    keymap[sap_field] = len(cols)
                    cols.append(c)
            continue
        if r == HEADER_ROW_DESC:
            descriptions = []
            for c in range(len(sap_fields)):
                long = _clean(_at(values, c))
                head_name = long.split("\n", 1)[0].rstrip("*").strip() if long else ""
                descriptions.append(head_name)
            continue
        if r < DATA_START_ROW:
            continue

        if not any(_clean(v) for v in values):
            continue

        vals: list[Any] = []
        ctypes: list[int] | None = [] if types is not None else None
        for c in cols:
            value = _at(values, c)
            if types is not None:
                ctype = types[c] if c < len(types) else xlrd.XL_CELL_EMPTY
                ctypes.append(ctype)
                # xlrd reads every numeric cell as a float, so an integer-valued
                # material number / plant / code (e.g. 1054) becomes 1054.0 and
                # renders with a trailing ".0". Normalise whole numbers to int.
                # Guard on XL_CELL_NUMBER only — dates are floats too and must
                # not be truncated; genuine decimals (12.5) are left untouched.
                if (ctype == xlrd.XL_CELL_NUMBER
                        and isinstance(value, float) and value.is_integer()):
                    value = int(value)
            else:
                # openpyxl: dates arrive as datetime objects, so any
                # integer-valued float here really is a number.
                if isinstance(value, float) and value.is_integer():
                    value = int(value)
            vals.append(value)

        rows.append(Row(
            row_num=r + 1,  # 1-based excel row number
            vals=tuple(vals),
            types=tuple(ctypes) if ctypes is not None else None,
            keymap=keymap,
        ))

    return SheetData(
        sheet=sheet_name,
        sap_structure=sap_structure,
        sap_fields=sap_fields,
        descriptions=descriptions,
        rows=rows,
    )


def open_workbook(file_path_or_bytes, file_name: str | None = None) -> _Book:
    """Open .xls or .xlsx — returns _Book wrapper."""
    if isinstance(file_path_or_bytes, (bytes, bytearray)):
        data = bytes(file_path_or_bytes)
        if data[:4] == b"PK\x03\x04":  # xlsx is a zip
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
            return _Book(wb, "openpyxl")
        return _Book(xlrd.open_workbook(file_contents=data), "xlrd")
    if isinstance(file_path_or_bytes, io.IOBase):
        data = file_path_or_bytes.read()
        return open_workbook(data, file_name)
    path = str(file_path_or_bytes)
    if path.lower().endswith(".xlsx"):
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True, read_only=True)
        return _Book(wb, "openpyxl")
    return _Book(xlrd.open_workbook(path), "xlrd")


def load_all(book: _Book) -> tuple[dict[tuple[str, str], FieldSpec], dict[str, SheetData]]:
    specs = load_field_list(book)
    sheets_in_specs = sorted({s for s, _ in specs.keys()})

    # If Field List is absent/empty, fall back to loading every sheet in the workbook
    if not sheets_in_specs:
        sheets_in_specs = book.sheet_names()

    data: dict[str, SheetData] = {}
    for s in sheets_in_specs:
        sd = load_sheet_data(book, s)
        if sd is not None:
            data[s] = sd
    book.close()
    return specs, data
