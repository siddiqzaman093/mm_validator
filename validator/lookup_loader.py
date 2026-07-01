"""
Load the Product Master Lookup File.

Sheets consumed:
  Field Types       — authoritative SAP type/length/decimal per field.
  Fields Entry      — per-product-type matrix: Material Class + Valuation Type.
  Plant-ProfitCenter — plant → expected profit center mapping.

Field Types sheet layout (merged-cell-style, forward-filled):
  Col 0  Sheet Name
  Col 1  Group Name
  Col 2  Field Description
  Col 3  Type (Text / Number / Date)
  Col 4  Length
  Col 5  Decimal
  Col 6  SAP Type    ← Priority 1 for type validation
  Col 7  SAP Length  ← Priority 1 for length validation
  Col 8  SAP Decimal ← Priority 1 for decimal validation

Fields Entry sheet layout (3-row header, then one row per product type):
  Row 1: Section headers (Sheet Name)
  Row 2: Group headers   (Group Name)
  Row 3: Field headers   (Material Type | Material Class | Valuation Type | ...)
  Row 4+: Data — one Z-code per row

Plant-ProfitCenter sheet layout:
  Row 1: Headers (Plant | Profit Center)
  Row 2+: Data
"""
from __future__ import annotations

import io
from dataclasses import dataclass


@dataclass
class LookupFieldSpec:
    sheet: str
    group: str
    description: str
    type: str | None        # UI type from lookup
    length: int | None      # UI length
    decimal: int | None     # UI decimal places
    sap_type: str | None    # SAP-native type  ← Priority 1 for type validation
    sap_length: int | None  # SAP field length ← Priority 1 for length validation
    sap_decimal: int | None # SAP decimal      ← Priority 1 for decimal validation


def _to_int(v) -> int | None:
    if v is None:
        return None
    try:
        f = float(v)
        return None if f != f else int(f)   # NaN guard
    except (TypeError, ValueError):
        return None


def _clean(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in ("nan", "none", "") else s


def load_lookup_field_types(file_bytes: bytes) -> dict[tuple[str, str], LookupFieldSpec]:
    """
    Parse the 'Field Types' sheet from the lookup workbook.

    Returns a dict keyed by (sheet_name_lower, field_description_lower).
    If the sheet is absent or the file is unreadable, returns an empty dict.
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
    except Exception:
        return {}

    if "Field Types" not in wb.sheetnames:
        return {}

    ws = wb["Field Types"]
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 2:
        return {}

    specs: dict[tuple[str, str], LookupFieldSpec] = {}
    current_sheet = ""
    current_group = ""

    for raw_row in rows[1:]:          # skip header row
        def _col(i):
            return raw_row[i] if i < len(raw_row) else None

        sheet_val = _clean(_col(0))
        group_val = _clean(_col(1))
        desc_val  = _clean(_col(2))
        type_val  = _clean(_col(3))
        len_val   = _to_int(_col(4))
        dec_val   = _to_int(_col(5))
        sap_type  = _clean(_col(6))
        sap_len   = _to_int(_col(7))
        sap_dec   = _to_int(_col(8))

        # Forward-fill merged cells
        if sheet_val:
            current_sheet = sheet_val
        if group_val:
            current_group = group_val

        if not desc_val or not current_sheet:
            continue

        key = (current_sheet.lower(), desc_val.lower())
        specs[key] = LookupFieldSpec(
            sheet=current_sheet,
            group=current_group,
            description=desc_val,
            type=type_val or None,
            length=len_val,
            decimal=dec_val,
            sap_type=sap_type or None,
            sap_length=sap_len,
            sap_decimal=sap_dec,
        )

    return specs


# ---------------------------------------------------------------------------
# Fields Entry — product-type matrix
# ---------------------------------------------------------------------------
def _norm_num(s: str) -> str:
    """Normalise a numeric string: '1000.0' → '1000', 'ABC' → 'ABC'."""
    try:
        return str(int(float(s)))
    except (ValueError, TypeError):
        return s


def load_lookup_fields_entry(file_bytes: bytes) -> dict[str, dict]:
    """
    Parse the 'Fields Entry' sheet.

    Returns  dict: material_type_upper → {
        'material_class': str | None,
        'valuation_type': str | None,
    }

    Sheet structure (3-row header):
      Row 3 (index 2) — field names: 'Material Type', 'Material Class',
                        'Valuation Type', and 643+ more columns.
      Rows 4+ (index 3+) — one Z-code per row.
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
    except Exception:
        return {}

    if "Fields Entry" not in wb.sheetnames:
        return {}

    ws = wb["Fields Entry"]
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 4:
        return {}

    # Row index 2 is the true field-header row
    header_row = rows[2]

    mat_type_col = mat_class_col = val_type_col = None
    for i, cell in enumerate(header_row):
        if cell is None:
            continue
        label = str(cell).strip().lower()
        if label == "material type" and mat_type_col is None:
            mat_type_col = i
        elif label == "material class" and mat_class_col is None:
            mat_class_col = i
        elif label == "valuation type" and val_type_col is None:
            # Take first occurrence (lives in Basic Data section)
            val_type_col = i

    if mat_type_col is None:
        return {}

    result: dict[str, dict] = {}
    for row in rows[3:]:           # data starts at index 3
        def _g(idx):
            return _clean(row[idx]) if idx is not None and idx < len(row) else ""

        mat_type = _g(mat_type_col)
        if not mat_type:
            continue
        result[mat_type.upper()] = {
            "material_class": _g(mat_class_col) or None,
            "valuation_type": _g(val_type_col)  or None,
        }

    return result


# ---------------------------------------------------------------------------
# Fields Entry — mandatory-field matrix
# ---------------------------------------------------------------------------
def load_lookup_mandatory_fields(file_bytes: bytes) -> dict[str, set[tuple[str, str]]]:
    """
    Build the per-material-type mandatory-field lookup from 'Fields Entry'.

    Returns
    -------
    dict: material_type_upper  →  set of (sheet_name_lower, field_desc_lower)
          for every field column that carries the value 'X' for that type.

    Sheet layout
    ------------
    Row 1 (idx 0): sparse section/sheet-name headers  → forward-fill per column
    Row 2 (idx 1): group sub-headers                  → not used here
    Row 3 (idx 2): field names                        → keyed per column
    Row 4+ (idx 3+): one data row per material type

    Columns 0-2 : Material Type | Material Class | Valuation Type  (metadata, skipped)
    Column  3+  : actual product-master fields  (user-facing "column 4 onwards")
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
    except Exception:
        return {}

    if "Fields Entry" not in wb.sheetnames:
        return {}

    ws = wb["Fields Entry"]
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 4:
        return {}

    section_row = rows[0]   # row 1 — sparse section/sheet names
    field_row   = rows[2]   # row 3 — field names

    FIELD_START_COL = 3     # 0-based; columns 0-2 are metadata

    # Forward-fill section names across all columns
    sections: list[str] = []
    current = ""
    for cell in section_row:
        v = _clean(cell)
        if v and v.lower() != "sheet name":   # skip the header label itself
            current = v
        sections.append(current)

    # Build col-index → (section_lower, field_desc_lower) for field columns only
    col_map: dict[int, tuple[str, str]] = {}
    for i, cell in enumerate(field_row):
        if i < FIELD_START_COL:
            continue
        field   = _clean(cell)
        section = sections[i] if i < len(sections) else ""
        if field and section:
            col_map[i] = (section.lower(), field.lower())

    # For each material-type row collect the set of (section, field) marked 'X'
    result: dict[str, set[tuple[str, str]]] = {}
    for row in rows[3:]:
        if not row or (len(row) > 0 and row[0] is None):
            continue
        mat_type = _clean(row[0]).upper()
        if not mat_type:
            continue

        mandatory: set[tuple[str, str]] = set()
        for col_idx, key in col_map.items():
            if col_idx < len(row) and _clean(row[col_idx]).upper() == "X":
                mandatory.add(key)

        result[mat_type] = mandatory

    return result


# ---------------------------------------------------------------------------
# Plant-ProfitCenter lookup
# ---------------------------------------------------------------------------
def load_lookup_plant_profit_center(file_bytes: bytes) -> dict[str, str]:
    """
    Parse the 'Plant-ProfitCenter' sheet.

    Returns  dict: normalised_plant_str → normalised_profit_center_str
    Numeric values (e.g. 1000.0) are normalised to their integer string form.
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
    except Exception:
        return {}

    if "Plant-ProfitCenter" not in wb.sheetnames:
        return {}

    ws = wb["Plant-ProfitCenter"]
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 2:
        return {}

    result: dict[str, str] = {}
    for row in rows[1:]:           # skip header
        if len(row) < 2:
            continue
        plant = _norm_num(_clean(row[0]))
        pc    = _norm_num(_clean(row[1]))
        if plant and pc:
            result[plant] = pc

    return result
